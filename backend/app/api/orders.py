from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional
import random, string, io, csv
from datetime import date as date_type

from app.core.database import get_db
from app.core.dependencies import get_current_user, get_current_admin
from app.models.user import User
from app.models.order import Order, OrderItem, OrderStatus
from app.models.product import ProductVariant
from app.models.category import Coupon
from app.schemas.order import (
    OrderCreate, OrderResponse, OrderListResponse, OrderStatusUpdate, TrackingUpdate
)
from app.utils.email import send_order_confirmation_email, send_tracking_email
from app.schemas.common import MessageResponse

router = APIRouter(prefix="/orders", tags=["orders"])

SHIPPING_COST = {"standard": 0.0, "express": 25.0}
TAX_RATE = 0.08


def _generate_order_number() -> str:
    return "AD-" + "".join(random.choices(string.digits, k=4))


@router.post("", response_model=OrderResponse, status_code=201)
async def create_order(
    data: OrderCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user),
):
    items_data = []
    subtotal = 0.0
    for item in data.items:
        variant = db.query(ProductVariant).filter(
            ProductVariant.id == item.variant_id
        ).first()
        if not variant:
            raise HTTPException(status_code=404, detail=f"Variante {item.variant_id} no encontrada")
        if variant.stock < item.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"{variant.product.name} ({variant.size_ml}ml): stock insuficiente",
            )
        total_price = variant.price * item.quantity
        subtotal += total_price
        items_data.append({
            "variant": variant,
            "quantity": item.quantity,
            "unit_price": variant.price,
            "total_price": total_price,
        })

    discount = 0.0
    if data.coupon_code:
        coupon = db.query(Coupon).filter(
            Coupon.code == data.coupon_code.upper(),
            Coupon.is_active == True,
        ).first()
        if not coupon:
            raise HTTPException(status_code=400, detail="Cupón inválido o expirado")
        if subtotal < coupon.min_order_amount:
            raise HTTPException(
                status_code=400,
                detail=f"El pedido mínimo para este cupón es ${coupon.min_order_amount}",
            )
        if coupon.discount_type == "percentage":
            discount = subtotal * (coupon.discount_value / 100)
        else:
            discount = coupon.discount_value
        coupon.used_count += 1

    shipping_cost = SHIPPING_COST.get(data.shipping_method, 0.0)
    tax = (subtotal - discount) * TAX_RATE
    total = subtotal - discount + shipping_cost + tax

    order_number = _generate_order_number()
    while db.query(Order).filter(Order.order_number == order_number).first():
        order_number = _generate_order_number()

    order = Order(
        order_number=order_number,
        user_id=current_user.id if current_user else None,
        subtotal=subtotal,
        shipping_cost=shipping_cost,
        tax=round(tax, 2),
        discount=round(discount, 2),
        total=round(total, 2),
        shipping_method=data.shipping_method,
        coupon_code=data.coupon_code,
        payment_method=data.payment_method,
        shipping_email=data.shipping.email,
        shipping_first_name=data.shipping.first_name,
        shipping_last_name=data.shipping.last_name,
        shipping_address=data.shipping.address,
        shipping_city=data.shipping.city,
        shipping_postal_code=data.shipping.postal_code,
        shipping_country=data.shipping.country,
    )
    db.add(order)
    db.flush()

    for item in items_data:
        variant = item["variant"]
        order_item = OrderItem(
            order_id=order.id,
            product_id=variant.product_id,
            variant_id=variant.id,
            product_name=variant.product.name,
            size_ml=variant.size_ml,
            quantity=item["quantity"],
            unit_price=item["unit_price"],
            total_price=item["total_price"],
        )
        variant.stock -= item["quantity"]
        db.add(order_item)

    db.commit()
    db.refresh(order)
    items_for_email = [
        {"product_name": i.product_name, "size_ml": i.size_ml, "quantity": i.quantity, "total_price": i.total_price}
        for i in order.items
    ]
    shipping_addr = f"{order.shipping_address}\n{order.shipping_city}, {order.shipping_postal_code}\n{order.shipping_country}"
    background_tasks.add_task(
        send_order_confirmation_email,
        order.shipping_email,
        order.shipping_first_name,
        order.order_number,
        items_for_email,
        order.total,
        shipping_addr,
    )
    return OrderResponse.model_validate(order)


@router.get("/my-orders", response_model=OrderListResponse)
async def my_orders(
    page: int = Query(1, ge=1),
    per_page: int = Query(10),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(Order).options(joinedload(Order.items)).filter(
        Order.user_id == current_user.id
    ).order_by(Order.created_at.desc())
    total = query.count()
    orders = query.offset((page - 1) * per_page).limit(per_page).all()
    return OrderListResponse(
        items=[OrderResponse.model_validate(o) for o in orders],
        total=total, page=page, per_page=per_page,
        pages=(total + per_page - 1) // per_page,
    )


@router.get("/export")
async def export_orders(
    format: str = Query("csv", regex="^(csv|excel)$"),
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
):
    from datetime import datetime as dt
    query = db.query(Order).options(joinedload(Order.items)).order_by(Order.created_at.desc())
    if status:
        query = query.filter(Order.status == status)
    if date_from:
        try:
            query = query.filter(Order.created_at >= dt.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Order.created_at <= dt.fromisoformat(date_to))
        except ValueError:
            pass
    orders = query.all()

    headers = [
        "Orden", "Fecha", "Cliente", "Email", "Productos",
        "Subtotal", "Impuestos", "Envío", "Descuento", "Total",
        "Método Pago", "Estado", "Tracking", "Empresa Envío", "Dirección"
    ]
    rows = []
    for o in orders:
        products_str = "; ".join([f"{i.product_name} ({i.size_ml}ml) x{i.quantity}" for i in o.items])
        rows.append([
            o.order_number,
            o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
            f"{o.shipping_first_name} {o.shipping_last_name}",
            o.shipping_email,
            products_str,
            o.subtotal,
            o.tax,
            o.shipping_cost,
            o.discount,
            o.total,
            o.payment_method or "",
            o.status.value if o.status else "",
            o.tracking_number or "",
            o.tracking_company or "",
            f"{o.shipping_address}, {o.shipping_city}, {o.shipping_postal_code}, {o.shipping_country}",
        ])

    filename = f"ordenes_aroma_{date_type.today().strftime('%Y%m%d')}"

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )
    else:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Órdenes"
        gold_fill = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
        bold_black = Font(bold=True, color="000000")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = gold_fill
            cell.font = bold_black
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )


@router.get("/{order_id}", response_model=OrderResponse)
async def get_order(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    order = db.query(Order).options(joinedload(Order.items)).filter(
        Order.id == order_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    if not current_user.is_admin and order.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Sin permisos")
    return OrderResponse.model_validate(order)


# Admin endpoints
@router.get("", response_model=OrderListResponse)
async def admin_list_orders(
    page: int = Query(1, ge=1),
    per_page: int = Query(20),
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
):
    query = db.query(Order).options(joinedload(Order.items)).order_by(Order.created_at.desc())
    if status:
        query = query.filter(Order.status == status)
    total = query.count()
    orders = query.offset((page - 1) * per_page).limit(per_page).all()
    return OrderListResponse(
        items=[OrderResponse.model_validate(o) for o in orders],
        total=total, page=page, per_page=per_page,
        pages=(total + per_page - 1) // per_page,
    )


@router.patch("/{order_id}/status", response_model=OrderResponse)
async def update_order_status(
    order_id: int,
    data: OrderStatusUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    order.status = data.status
    db.commit()
    db.refresh(order)
    return OrderResponse.model_validate(order)


@router.get("/{order_id}/confirmation", response_model=OrderResponse)
async def get_order_confirmation(
    order_id: int,
    db: Session = Depends(get_db),
):
    order = db.query(Order).options(joinedload(Order.items)).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    return OrderResponse.model_validate(order)


@router.put("/{order_id}/tracking", response_model=OrderResponse)
async def update_tracking(
    order_id: int,
    data: TrackingUpdate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
):
    order = db.query(Order).options(joinedload(Order.items)).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    order.tracking_number = data.tracking_number
    order.tracking_company = data.tracking_company
    order.tracking_url = data.tracking_url
    order.status = OrderStatus.shipped
    db.commit()
    db.refresh(order)
    if order.shipping_email:
        background_tasks.add_task(
            send_tracking_email,
            order.shipping_email,
            order.shipping_first_name,
            order.order_number,
            data.tracking_number,
            data.tracking_company,
            data.tracking_url,
        )
    return OrderResponse.model_validate(order)
